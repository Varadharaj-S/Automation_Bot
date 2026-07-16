"""
routes/tracker.py — the problem list/search page, xlsx export, the Daily
Tracker feature (admin-defined auto-evaluated sheets), and the per-student
tracker sheet view. Moved verbatim from app.py.
"""

import io
import json
from datetime import datetime
from collections import Counter

from flask import render_template, request, send_file, jsonify, redirect, url_for, flash, abort
from flask_login import current_user

from extensions import app
from database.db import get_db
from utils.decorators import login_required, verified_required, admin_required, log_admin
from utils.security import sanitize
from services.tracker_service import ensure_tracker_schema, refresh_tracker_sheet_results
from utils.helpers import _get_user_by_username, _group_rows, _parse_any_date


# ── Problem List ──────────────────────────────────────────────────────────────
@app.route("/problems")
@login_required
@verified_required
def problems():
    platform = request.args.get("platform", "all")
    diff     = request.args.get("diff", "all")
    search   = sanitize(request.args.get("q", ""), 64)
    page     = max(1, int(request.args.get("page", 1)))
    per_page = 50

    with get_db() as db:
        base   = "SELECT * FROM submissions WHERE user_id=?"
        params = [current_user.id]
        if platform != "all":
            base += " AND platform=?"; params.append(platform)
        if diff != "all":
            base += " AND difficulty=?"; params.append(diff)
        if search:
            base += " AND (problem_name ILIKE ? OR tags ILIKE ?)"; params += [f"%{search}%", f"%{search}%"]
        total = db.execute(f"SELECT COUNT(*) as c FROM ({base}) AS sub", params).fetchone()["c"]
        base += " ORDER BY solved_date DESC LIMIT ? OFFSET ?"
        params += [per_page, (page - 1) * per_page]
        subs   = db.execute(base, params).fetchall()

    return render_template("problems.html",
                           subs=subs, total=total, page=page,
                           total_pages=(total + per_page - 1) // per_page,
                           platform=platform, diff=diff, search=search)


# ── Export XLSX ───────────────────────────────────────────────────────────────
@app.route("/export_csv")
@login_required
@verified_required
def export_csv():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    with get_db() as db:
        subs = db.execute(
            "SELECT solved_date,problem_name,problem_url,difficulty,platform,tags "
            "FROM submissions WHERE user_id=? ORDER BY solved_date DESC",
            (current_user.id,)
        ).fetchall()
    topic_counter: Counter = Counter()
    for s in subs:
        tags = s["tags"] or ""
        first = next((t.strip() for t in tags.split(";") if t.strip()), "General")
        topic_counter[first] += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{current_user.username} DSA"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    headers = ["DATE", "PROGRAM TITLE", "LINK", "DIFFICULTY", "PLATFORM", "TOPIC", "COUNT"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    diff_colors = {"Easy": "C6EFCE", "Medium": "FFEB9C", "Hard": "FFC7CE"}
    for s in subs:
        tags  = s["tags"] or ""
        first = next((t.strip() for t in tags.split(";") if t.strip()), "General")
        name  = str(s["problem_name"] or "").lstrip("=+-@")
        ws.append([s["solved_date"], name, s["problem_url"], s["difficulty"],
                   s["platform"], first, topic_counter[first]])
        diff = s["difficulty"] or ""
        color = diff_colors.get(diff)
        if color:
            ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=color)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"dsa_{current_user.username}.xlsx")


# ── Daily Tracker ─────────────────────────────────────────────────────────────
@app.route("/daily_tracker")
@login_required
@verified_required
def daily_tracker():
    """
    Master daily tracker page.
    Admin sees all students; regular users are redirected to their own sheet.
    """
    if not current_user.is_admin:
        return redirect(url_for("student_sheet", username=current_user.username))

    ensure_tracker_schema()
    refresh_tracker_sheet_results()

    with get_db() as db:
        users = db.execute("""
            SELECT id, username FROM users
            WHERE status='active' AND is_admin=0
            ORDER BY username ASC
        """).fetchall()
        sheets = db.execute("""
            SELECT * FROM daily_tracker_sheets
            ORDER BY id ASC
        """).fetchall()
        results = db.execute("""
            SELECT r.*, s.name as sheet_name, u.username
            FROM daily_tracker_sheet_results r
            JOIN daily_tracker_sheets s ON s.id=r.sheet_id
            JOIN users u ON u.id=r.student_id
            ORDER BY r.updated_at DESC, u.username ASC
        """).fetchall()
        subs = db.execute("""
            SELECT s.*, u.username
            FROM submissions s
            JOIN users u ON u.id=s.user_id
            WHERE u.status='active' AND u.is_admin=0
            ORDER BY s.solved_date DESC, u.username ASC, s.id DESC
        """).fetchall()

    daily_rows = _group_rows(subs, include_username=True)

    today_label_iso = datetime.now().strftime("%Y-%m-%d")
    today_label_dmy = datetime.now().strftime("%d-%m-%Y")  # legacy compat
    active_today = len({row["username"] for row in daily_rows
                        if row["date"] in (today_label_iso, today_label_dmy)})

    summary_cards = {
        "students": len(users),
        "active_today": active_today,
        "total_solved": len(subs),
        "sheets": len(sheets)
    }

    return render_template(
        "daily_tracker.html",
        is_admin=True,
        view_username=None,
        sheets=sheets,
        results=results,
        daily_rows=daily_rows,
        summary_cards=summary_cards,
        users=users,
        page_title="Master Daily Tracker"
    )


@app.route("/student_sheet/<username>")
@login_required
@verified_required
def student_sheet(username):
    ensure_tracker_schema()
    target = _get_user_by_username(username)
    if not target:
        abort(404)

    self_view = current_user.id == target["id"]
    admin_view = bool(current_user.is_admin)

    if not self_view and not admin_view:
        abort(403)

    refresh_tracker_sheet_results()

    with get_db() as db:
        subs = db.execute("""
            SELECT * FROM submissions
            WHERE user_id=%s
            ORDER BY solved_date DESC, id DESC
        """, (target["id"],)).fetchall()

        sheets = db.execute("""
            SELECT * FROM daily_tracker_sheets
            ORDER BY id ASC
        """).fetchall()

        results = db.execute("""
            SELECT r.*, s.name as sheet_name
            FROM daily_tracker_sheet_results r
            JOIN daily_tracker_sheets s ON s.id=r.sheet_id
            WHERE r.student_id=?
            ORDER BY s.id ASC
        """, (target["id"],)).fetchall()

    daily_rows = _group_rows(subs, include_username=False)

    today_str = datetime.now().strftime("%Y-%m-%d")
    today_subs = [s for s in subs if (s["solved_date"] or "").startswith(today_str)]
    last_sub = subs[0] if subs else None

    # Find target for Daily Problems sheet
    daily_sheet = next((s for s in sheets if s["sheet_type"] == "daily_problems"), None)
    daily_target = int(daily_sheet["target_value"]) if daily_sheet and daily_sheet["target_value"] else 10

    summary_cards = {
        "total_solved": len(subs),
        "active_days": len(daily_rows),
        "current_streak": 0,
        "longest_streak": 0,
        "sheets": len(sheets),
        "today_solved": len(today_subs),
        "today_target": daily_target,
        "today_remaining": max(0, daily_target - len(today_subs)),
        "last_problem": last_sub["problem_name"] if last_sub else None,
        "last_solved_date": last_sub["solved_date"] if last_sub else None,
    }

    # simple streak calc from submission dates
    parsed_dates = sorted({(_parse_any_date(r["solved_date"]).date()) for r in subs if _parse_any_date(r["solved_date"])})
    if parsed_dates:
        # current streak = consecutive days ending with the most recent active day
        current = 1
        for prev, cur in zip(parsed_dates[-2::-1], parsed_dates[::-1]):
            if (cur - prev).days == 1:
                current += 1
            else:
                break

        longest = 1
        run = 1
        for prev, cur in zip(parsed_dates[:-1], parsed_dates[1:]):
            if (cur - prev).days == 1:
                run += 1
                longest = max(longest, run)
            else:
                run = 1

        summary_cards["current_streak"] = current
        summary_cards["longest_streak"] = longest

    return render_template(
        "daily_tracker.html",
        is_admin=False,
        view_username=target["username"],
        target_user=target,
        sheets=sheets,
        results=results,
        daily_rows=daily_rows,
        summary_cards=summary_cards,
        users=[],
        page_title=f"{target['username']} – My Sheet",
        self_view=self_view,
        admin_view=admin_view
    )


# ── Admin: create custom tracker sheet ─────────────────────────────────────────
@app.route("/admin/tracker_sheets/create", methods=["POST"])
@login_required
@admin_required
def create_tracker_sheet():
    ensure_tracker_schema()
    name = sanitize(request.form.get("name", ""), 80)
    sheet_type = sanitize(request.form.get("sheet_type", "custom"), 40).lower()
    description = sanitize(request.form.get("description", ""), 200)
    target_value = request.form.get("target_value", "0").strip()
    target_topic = sanitize(request.form.get("target_topic", ""), 80)
    visibility = sanitize(request.form.get("visibility", "private"), 20).lower()

    if not name:
        flash("Sheet name is required.", "error")
        return redirect(url_for("daily_tracker"))

    try:
        target_value = int(target_value or 0)
    except Exception:
        target_value = 0

    with get_db() as db:
        try:
            db.execute("""
                INSERT INTO daily_tracker_sheets
                (name, sheet_type, description, target_value, target_topic, created_by, visibility, created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (
                name, sheet_type, description, target_value, target_topic,
                current_user.id, visibility, datetime.now().isoformat(timespec="seconds")
            ))
            db.commit()
            flash(f"Created sheet '{name}'.", "success")
        except Exception as e:
            flash(f"Could not create sheet: {e}", "error")

    return redirect(url_for("daily_tracker"))


@app.route("/admin/tracker_results/create", methods=["POST"])
@login_required
@admin_required
def create_tracker_result():
    ensure_tracker_schema()
    sheet_id = request.form.get("sheet_id", "").strip()
    username = sanitize(request.form.get("username", ""), 64)
    status = sanitize(request.form.get("status", ""), 40)
    remarks = sanitize(request.form.get("remarks", ""), 240)
    solved_count = request.form.get("solved_count", "0").strip()

    try:
        sheet_id = int(sheet_id)
        solved_count = int(solved_count or 0)
    except Exception:
        flash("Invalid sheet or solved count.", "error")
        return redirect(url_for("daily_tracker"))

    user = _get_user_by_username(username)
    if not user:
        flash("Username not found.", "error")
        return redirect(url_for("daily_tracker"))

    with get_db() as db:
        db.execute("""
            INSERT INTO daily_tracker_sheet_results
            (sheet_id, student_id, status, solved_count, details_json, remarks, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet_id, student_id) DO UPDATE SET
                status=excluded.status,
                solved_count=excluded.solved_count,
                details_json=excluded.details_json,
                remarks=excluded.remarks,
                updated_at=excluded.updated_at
        """, (
            sheet_id, user["id"], status, solved_count,
            json.dumps({"manual": True}),
            remarks,
            datetime.now().isoformat(timespec="seconds")
        ))
        db.commit()
    flash(f"Updated result for {username}.", "success")
    return redirect(url_for("daily_tracker"))


@app.route("/admin/tracker_refresh", methods=["POST"])
@login_required
@admin_required
def refresh_tracker():
    ensure_tracker_schema()
    refresh_tracker_sheet_results()
    flash("Tracker results refreshed.", "success")
    return redirect(url_for("daily_tracker"))

