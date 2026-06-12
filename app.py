"""
DSA Tracker v3
──────────────
Fixes: rate limiter, secret key, DB patterns, auto-sync stability
New:   LeetCode password login, daily challenges, mentor mode,
       custom problems (admin), separate admin login,
       platform checkboxes, auto Google Sheet creation, admin analytics
"""

from flask import jsonify
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
except Exception:
    webdriver = None
    Service = None
    ChromeDriverManager = None
import time
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_file, abort, session)
from flask_login import (LoginManager, UserMixin, login_user,
                         logout_user, login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import subprocess
import sys
from datetime import datetime, timedelta
from collections import Counter
import sqlite3, os, io, csv, json, shutil, threading, time, re, secrets

try:
    from normal_sync import sync_user_data
except Exception:
    def sync_user_data(*args, **kwargs):
        return None

try:
    from bot import get_dashboard_data, generate_daily_challenges, create_user_sheet, import_lc_full
except Exception:
    def get_dashboard_data(*args, **kwargs):
        return {"chart_labels": [], "chart_data": [], "total": 0, "solved": 0}
    def generate_daily_challenges(*args, **kwargs):
        return []
    def create_user_sheet(*args, **kwargs):
        return None
    def import_lc_full(*args, **kwargs):
        return {"success": False, "message": "LeetCode import unavailable in this environment."}
import time
import subprocess
import threading
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
load_dotenv()

# ── Email Verification Token Store (in-memory + DB backed) ───────────────────
_verify_tokens = {}  # token -> user_id

def send_verification_email(to_email, username, token):
    """Send email verification link. Configure SMTP via env vars."""
    smtp_host   = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port   = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user   = os.environ.get("SMTP_USER", "")
    smtp_pass   = os.environ.get("SMTP_PASS", "")
    from_email  = os.environ.get("FROM_EMAIL", smtp_user)

    if not smtp_user or not smtp_pass:
        print(f"[Email] SMTP not configured — skip verification for {username}")
        return False

    verify_url = os.environ.get("APP_URL", "http://localhost:5000") + f"/verify_email/{token}"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = "DSA Tracker — Verify your email"
    msg["From"]    = from_email
    msg["To"]      = to_email

    html = f"""
    <div style="font-family:sans-serif;max-width:480px;margin:auto;padding:2rem;background:#0f172a;color:#e2e8f0;border-radius:12px">
      <h2 style="color:#4d9de0">⚡ DSA Tracker</h2>
      <p>Hi <strong>{username}</strong>! Verify your email to activate your account.</p>
      <a href="{verify_url}"
         style="display:inline-block;margin:1.5rem 0;padding:.75rem 2rem;
                background:#4d9de0;color:#fff;text-decoration:none;
                border-radius:8px;font-weight:700">
        ✅ Verify Email
      </a>
      <p style="color:#64748b;font-size:.85rem">
        Link expires in 24 hours. If you did not sign up, ignore this email.
      </p>
    </div>"""

    msg.attach(MIMEText(html, "html"))
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, to_email, msg.as_string())
        print(f"[Email] Verification sent to {to_email}")
        return True
    except Exception as e:
        print(f"[Email] Failed: {e}")
        return False

#-----------------------------------------------------------

import sqlite3

def get_counts(user_id):
    conn = sqlite3.connect("dsa_tracker.db")
    cursor = conn.cursor()

    # ✅ use submissions table
    cursor.execute("""
SELECT COUNT(DISTINCT problem_id)
FROM submissions
WHERE user_id=?
""", (user_id,))
    total = cursor.fetchone()[0] or 0

    # ✅ solved = all rows (already solved submissions)
    conn.close()
    return total, total



# ── App ───────────────────────────────────────────────────────────────────────
app = Flask(__name__)
# SECURITY: load from env, generate random fallback (never hardcoded)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=12)
app.config["SESSION_COOKIE_HTTPONLY"]    = True
app.config["SESSION_COOKIE_SAMESITE"]   = "Lax"

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Please log in to continue."

DB         = os.environ.get("DB_PATH", "dsa_tracker.db")
BACKUP_DIR = "backups"
os.makedirs(BACKUP_DIR, exist_ok=True)

# ── Rate Limiter (fixed — properly wrapped) ───────────────────────────────────
_rate_data = {}
_rate_lock = threading.Lock()

def is_rate_limited(key, max_calls=5, window=60):
    """Returns True if key has exceeded max_calls in window seconds."""
    now = time.time()
    with _rate_lock:
        calls = [t for t in _rate_data.get(key, []) if now - t < window]
        if len(calls) >= max_calls:
            return True
        calls.append(now)
        _rate_data[key] = calls
    return False

def rate_limit(max_calls=5, window=60):
    """Properly working rate-limit decorator."""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            key = f"{request.remote_addr}:{f.__name__}"
            if is_rate_limited(key, max_calls, window):
                if request.is_json or request.headers.get("X-Requested-With"):
                    return jsonify({"success": False,
                                    "message": "Too many requests. Please wait."}), 429
                flash("Too many requests. Please wait.", "error")
                return redirect(request.referrer or url_for("login"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


#-----User Settings-------------------------------------------------------------------------
from flask import request, jsonify
from flask_login import login_required, current_user

@app.route("/save_settings", methods=["POST"])
@login_required
def save_settings():
    data = request.get_json() or {}

    with get_db() as db:
        db.execute("""
        UPDATE users SET
            lc_handle=?,
            cf_handle=?,
            ac_handle=?,
            lc_session_cookie=?,
            lc_csrf_token=?
        WHERE id=?
        """, (
            data.get("leetcode", ""),
            data.get("codeforces", ""),
            data.get("atcoder", ""),
            data.get("cookie", ""),
            data.get("csrf", ""),
            current_user.id
        ))

        db.commit()

    return jsonify({"status": "saved"})

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            username         TEXT UNIQUE NOT NULL,
            email            TEXT DEFAULT '',
            password         TEXT NOT NULL,
            is_admin         INTEGER DEFAULT 0,
            is_verified      INTEGER DEFAULT 0,
            status           TEXT DEFAULT 'pending',
            sheet_id         TEXT DEFAULT '',
            cf_handle        TEXT DEFAULT '',
            lc_handle        TEXT DEFAULT '',
            lc_password      TEXT DEFAULT '',
            lc_session_cookie TEXT DEFAULT '',
            lc_csrf_token    TEXT DEFAULT '',
            ac_handle        TEXT DEFAULT '',
            enabled_platforms TEXT DEFAULT '["Codeforces","LeetCode","AtCoder"]',
            bio              TEXT DEFAULT '',
            is_public        INTEGER DEFAULT 1,
            last_login       TEXT DEFAULT '',
            last_sync        TEXT DEFAULT '',
            created_at       TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS submissions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL,
            platform     TEXT NOT NULL,
            problem_name TEXT,
            problem_id   TEXT,
            problem_url  TEXT DEFAULT '',
            difficulty   TEXT DEFAULT '',
            tags         TEXT DEFAULT '',
            solved_date  TEXT,
            UNIQUE(user_id, platform, problem_id)
        );
        CREATE INDEX IF NOT EXISTS idx_sub_user     ON submissions(user_id);
        CREATE INDEX IF NOT EXISTS idx_sub_platform ON submissions(user_id, platform);
        CREATE TABLE IF NOT EXISTS follows (
            follower_id  INTEGER NOT NULL,
            following_id INTEGER NOT NULL,
            created_at   TEXT DEFAULT '',
            PRIMARY KEY (follower_id, following_id)
        );
        CREATE TABLE IF NOT EXISTS admin_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id    INTEGER,
            action      TEXT,
            target_user TEXT,
            details     TEXT,
            ip          TEXT,
            created_at  TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS login_history (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER,
            ip         TEXT,
            user_agent TEXT,
            success    INTEGER,
            created_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS sync_logs (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER,
            status     TEXT,
            message    TEXT,
            created_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS daily_challenges (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL,
            problem_name  TEXT,
            problem_url   TEXT,
            difficulty    TEXT,
            platform      TEXT,
            topic         TEXT,
            assigned_date TEXT,
            completed     INTEGER DEFAULT 0,
            assigned_by   INTEGER DEFAULT 0,
            UNIQUE(user_id, problem_url, assigned_date)
        );
        CREATE TABLE IF NOT EXISTS mentor_assignments (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id      INTEGER,
            user_id       INTEGER,
            problem_name  TEXT,
            problem_url   TEXT,
            difficulty    TEXT,
            platform      TEXT,
            topic         TEXT,
            note          TEXT DEFAULT '',
            assigned_date TEXT,
            due_date      TEXT DEFAULT '',
            completed     INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS custom_problems (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            created_by  INTEGER,
            title       TEXT,
            description TEXT,
            difficulty  TEXT,
            topic       TEXT,
            url         TEXT DEFAULT '',
            created_at  TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS email_tokens (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            token      TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT '',
            used       INTEGER DEFAULT 0
        );
        """)
        # Default admin
        if not db.execute("SELECT id FROM users WHERE username='admin'").fetchone():
            db.execute("""
                INSERT INTO users
                (username,email,password,is_admin,is_verified,status,
                 enabled_platforms,created_at)
                VALUES (?,?,?,1,1,'active',?,?)
            """, ("admin","admin@dsatracker.local",
                  generate_password_hash("admin123"),
                  '["Codeforces","LeetCode","AtCoder"]',
                  datetime.now().isoformat()))
        db.commit()

init_db()
# ── User model ────────────────────────────────────────────────────────────────
class User(UserMixin):
    def __init__(self, row):
        for k in row.keys():
            setattr(self, k, row[k])
        for attr in ["sheet_id","cf_handle","lc_handle","lc_password","lc_session_cookie","lc_csrf_token",
                     "ac_handle","bio","email","last_login","last_sync","lc_imported","auto_sync_enabled","sync_time"]:
            if not getattr(self, attr, None):
                setattr(self, attr, "")
        if not getattr(self, "enabled_platforms", None):
            self.enabled_platforms = '["Codeforces","LeetCode","AtCoder"]'
        if not getattr(self, "lc_imported", None):
            self.lc_imported = 0
        if not getattr(self, "auto_sync_enabled", None):
            self.auto_sync_enabled = 1
        if not getattr(self, "sync_time", None):
            self.sync_time = "09:00"

    @property
    def enabled_list(self):
        try:
            return json.loads(self.enabled_platforms)
        except:
            return ["Codeforces","LeetCode","AtCoder"]

@login_manager.user_loader
def load_user(uid):
    with get_db() as db:
        row = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    return User(row) if row else None

# ── Decorators ────────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def wrapped(*a, **kw):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Admin access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*a, **kw)
    return wrapped

def verified_required(f):
    @wraps(f)
    def wrapped(*a, **kw):
        if not current_user.is_verified or current_user.status != "active":
            return redirect(url_for("pending"))
        return f(*a, **kw)
    return wrapped

def sanitize(s, maxlen=128):
    if not s: return ""
    return re.sub(r'[<>\'\";&]', '', str(s).strip()[:maxlen])

def log_admin(action, target="", details=""):
    with get_db() as db:
        db.execute("""
            INSERT INTO admin_logs
            (admin_id,action,target_user,details,ip,created_at)
            VALUES (?,?,?,?,?,?)
        """, (current_user.id, action, target, details,
              request.remote_addr, datetime.now().isoformat()))
        db.commit()
#-----------------------------------------------------
from datetime import datetime, timedelta
from flask import request

@app.route("/weekly_report")
@login_required
@verified_required
def weekly_report():

    start_param = request.args.get("start")
    # Admin can view all users combined OR a specific user
    view_all    = request.args.get("view") == "all" and current_user.is_admin
    target_uid  = request.args.get("uid")  # admin viewing a specific user

    if start_param:
        start = datetime.strptime(start_param, "%Y-%m-%d")
    else:
        start = datetime.now() - timedelta(days=6)

    end = start + timedelta(days=6)
    start_str = start.strftime("%Y-%m-%d")
    end_str   = end.strftime("%Y-%m-%d")

    # Determine which user(s) to query
    if view_all:
        # Admin: all active users combined
        query_uid  = None
        report_label = "All Users"
    elif target_uid and current_user.is_admin:
        query_uid  = int(target_uid)
        with get_db() as db:
            u = db.execute("SELECT username FROM users WHERE id=?", (query_uid,)).fetchone()
        report_label = u["username"] if u else "Unknown"
    else:
        query_uid  = current_user.id
        report_label = current_user.username

    with get_db() as db:
        if view_all:
            subs = db.execute("""
                SELECT s.*, u.username FROM submissions s
                JOIN users u ON s.user_id = u.id
                WHERE s.solved_date BETWEEN ? AND ?
                AND u.status='active' AND u.is_admin=0
            """, (start_str, end_str)).fetchall()
        else:
            subs = db.execute("""
                SELECT * FROM submissions
                WHERE user_id=? AND solved_date BETWEEN ? AND ?
            """, (query_uid, start_str, end_str)).fetchall()

    # ── Per-user breakdown for admin all-view ──────────────────────────────
    user_breakdown = {}
    if view_all:
        for s in subs:
            uname = s["username"]
            if uname not in user_breakdown:
                user_breakdown[uname] = {"solved": 0, "cf": 0, "lc": 0, "ac": 0}
            user_breakdown[uname]["solved"] += 1
            p = s["platform"]
            if p == "Codeforces":   user_breakdown[uname]["cf"] += 1
            elif p == "LeetCode":   user_breakdown[uname]["lc"] += 1
            elif p == "AtCoder":    user_breakdown[uname]["ac"] += 1
        # sort by solved desc
        user_breakdown = dict(sorted(user_breakdown.items(),
                                     key=lambda x: x[1]["solved"], reverse=True))

    # ── Main calculations ──────────────────────────────────────────────────
    solved_count = len(subs)
    cf = lc = ac = 0
    topic_count = {}
    days = set()

    for s in subs:
        p = s["platform"]
        if p == "Codeforces":   cf += 1
        elif p == "LeetCode":   lc += 1
        elif p == "AtCoder":    ac += 1

        tags  = s["tags"] or ""
        topic = tags.split(";")[0] if tags else "General"
        topic_count[topic] = topic_count.get(topic, 0) + 1
        days.add(s["solved_date"])

    if len(topic_count) >= 2:
        top_topic  = max(topic_count, key=topic_count.get)
        weak_topic = min(topic_count, key=topic_count.get)
        # Ensure strong != weak (happens when all topics have equal count)
        if top_topic == weak_topic:
            weak_topic = "Insufficient data"
    elif len(topic_count) == 1:
        top_topic  = list(topic_count.keys())[0]
        weak_topic = "Insufficient data"
    else:
        top_topic  = "N/A"
        weak_topic = "N/A"
    streak     = len(days)

    report = {
        "solved_count": solved_count,
        "cf_count":     cf,
        "lc_count":     lc,
        "ac_count":     ac,
        "top_topic":    top_topic,
        "weak_topic":   weak_topic,
        "streak":       streak,
        "week_start":   start_str,
        "label":        report_label
    }

    # ── History (last 5 weeks) ─────────────────────────────────────────────
    history = []
    with get_db() as db:
        for i in range(1, 6):
            s = start - timedelta(days=7*i)
            e = s + timedelta(days=6)
            s_str = s.strftime("%Y-%m-%d")
            e_str = e.strftime("%Y-%m-%d")

            if view_all:
                rows = db.execute("""
                    SELECT s.* FROM submissions s
                    JOIN users u ON s.user_id=u.id
                    WHERE s.solved_date BETWEEN ? AND ?
                    AND u.status='active' AND u.is_admin=0
                """, (s_str, e_str)).fetchall()
            else:
                rows = db.execute("""
                    SELECT * FROM submissions
                    WHERE user_id=? AND solved_date BETWEEN ? AND ?
                """, (query_uid, s_str, e_str)).fetchall()

            if not rows:
                continue

            h_cf = h_lc = h_ac = 0
            for r in rows:
                if r["platform"] == "Codeforces":   h_cf += 1
                elif r["platform"] == "LeetCode":   h_lc += 1
                elif r["platform"] == "AtCoder":    h_ac += 1

            history.append({
                "week_start":   s_str,
                "solved_count": len(rows),
                "cf_count":     h_cf,
                "lc_count":     h_lc,
                "ac_count":     h_ac
            })

    # Admin: list of all users for dropdown
    all_users = []
    if current_user.is_admin:
        with get_db() as db:
            all_users = db.execute(
                "SELECT id, username FROM users WHERE status='active' AND is_admin=0 ORDER BY username"
            ).fetchall()

    return render_template(
        "weekly_report.html",
        report=report,
        history=history,
        user=current_user,
        view_all=view_all,
        user_breakdown=user_breakdown,
        all_users=all_users,
        target_uid=target_uid
    )

#------------weekly report csv -----------------------------

@app.route("/weekly_csv")
@login_required
def weekly_csv():
    today = datetime.now()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)

    start_str = start.strftime("%Y-%m-%d")
    end_str = end.strftime("%Y-%m-%d")

    with get_db() as db:
        subs = db.execute("""
            SELECT solved_date,problem_name,problem_url,difficulty,platform,tags
            FROM submissions
            WHERE user_id=? AND solved_date BETWEEN ? AND ?
        """, (current_user.id, start_str, end_str)).fetchall()

    out = io.StringIO()
    w = csv.writer(out)

    w.writerow(["DATE","NAME","LINK","DIFFICULTY","PLATFORM","TOPIC"])

    for s in subs:
        topic = (s["tags"] or "").split(";")[0] if s["tags"] else "General"
        w.writerow([
            s["solved_date"],
            s["problem_name"],
            s["problem_url"],
            s["difficulty"],
            s["platform"],
            topic
        ])

    out.seek(0)

    return send_file(
        io.BytesIO(out.read().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name="weekly_report.csv"
    )

@app.route("/api/weekly_report")
@login_required
def api_weekly_report():
    import subprocess

    user_id = current_user.id

    subprocess.Popen(
        ["python", "bot_sheet_sync.py", str(user_id)]
    )

    return {"status": "started"}
# ── Index ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for("admin_dashboard"))
        if not current_user.is_verified or current_user.status != "active":
            return redirect(url_for("pending"))
        return redirect(url_for("dashboard"))  # ✅ correct now
    return redirect(url_for("login"))

# ── User Auth ─────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
@rate_limit(max_calls=10, window=60)
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = sanitize(request.form.get("username",""))
        password = request.form.get("password","")
        ip, ua   = request.remote_addr, request.user_agent.string[:200]

        with get_db() as db:
            row = db.execute(
                "SELECT * FROM users WHERE username=? AND is_admin=0", (username,)
            ).fetchone()
            ok = bool(row and check_password_hash(row["password"], password))
            db.execute("""
                INSERT INTO login_history
                (user_id,ip,user_agent,success,created_at)
                VALUES (?,?,?,?,?)
            """, (row["id"] if row else None, ip, ua,
                  1 if ok else 0, datetime.now().isoformat()))
            if ok:
                db.execute("UPDATE users SET last_login=? WHERE id=?",
                           (datetime.now().strftime("%Y-%m-%d %H:%M"), row["id"]))
            db.commit()

        if ok:
            user = User(row)
            if row["status"] != "active":
                flash("Your account is pending admin verification.", "warning")
                return redirect(url_for("pending"))
            login_user(user, remember=True)
            flash(f"Welcome back, {username}!", "success")
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")

@app.route("/signup", methods=["GET","POST"])
@rate_limit(max_calls=5, window=300)
def signup():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = sanitize(request.form.get("username",""), 32)
        email    = sanitize(request.form.get("email",""), 120)
        password = request.form.get("password","")
        confirm  = request.form.get("confirm","")

        if not username or len(username) < 3:
            flash("Username must be at least 3 characters.", "error")
        elif not re.match(r'^[a-zA-Z0-9_]+$', username):
            flash("Username: letters, numbers, _ only.", "error")
        elif password != confirm:
            flash("Passwords do not match.", "error")
        elif len(password) < 6:
            flash("Password min. 6 characters.", "error")
        else:
            try:
                with get_db() as db:
                    db.execute("""
                        INSERT INTO users
                        (username,email,password,is_verified,status,
                         enabled_platforms,created_at)
                        VALUES (?,?,?,0,'pending',?,?)
                    """, (username, email, generate_password_hash(password),
                          '["Codeforces","LeetCode","AtCoder"]',
                          datetime.now().isoformat()))
                    db.commit()
                    user_row = db.execute(
                        "SELECT id FROM users WHERE username=?", (username,)
                    ).fetchone()
                    uid = user_row["id"]
                    token = secrets.token_urlsafe(32)
                    db.execute("""
                        INSERT INTO email_tokens (user_id, token, created_at)
                        VALUES (?,?,?)
                    """, (uid, token, datetime.now().isoformat()))
                    db.commit()

                threading.Thread(
                    target=send_verification_email,
                    args=(email, username, token),
                    daemon=True
                ).start()

                flash("Account created! Please check your email to verify your address, then wait for admin approval.", "success")
                return redirect(url_for("pending"))
            except sqlite3.IntegrityError:
                flash("Username already taken.", "error")
    return render_template("signup.html")

def _create_sheet_bg(username, email):
    """Background: create Google Sheet and save ID to user record."""
    sheet_id, msg = create_user_sheet(email, username)
    if sheet_id:
        with get_db() as db:
            db.execute("UPDATE users SET sheet_id=? WHERE username=?",
                       (sheet_id, username))
            db.commit()
        print(f"[Sheets] Created sheet for {username}: {sheet_id}")
    else:
        print(f"[Sheets] Failed for {username}: {msg}")

@app.route("/verify_email/<token>")
def verify_email(token):
    with get_db() as db:
        row = db.execute(
            "SELECT * FROM email_tokens WHERE token=? AND used=0", (token,)
        ).fetchone()
        if not row:
            flash("Invalid or expired verification link.", "error")
            return redirect(url_for("login"))

        # Check token not older than 24h
        created = datetime.fromisoformat(row["created_at"])
        if datetime.now() - created > timedelta(hours=24):
            flash("Verification link has expired. Please contact admin.", "error")
            return redirect(url_for("login"))

        uid = row["user_id"]
        db.execute("UPDATE email_tokens SET used=1 WHERE token=?", (token,))
        db.execute("UPDATE users SET is_verified=1 WHERE id=? AND status='pending'", (uid,))
        # Note: status stays 'pending' until admin approves — is_verified=1 just means email confirmed
        db.commit()

        user_row = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
        username = user_row["username"] if user_row else "User"

    flash(f"✅ Email verified! Hi {username} — your account is now pending admin approval.", "success")
    return redirect(url_for("pending"))

@app.route("/pending")
def pending():
    return render_template("pending.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out successfully.", "success")
    return redirect(url_for("login"))

# ── Separate Admin Login ───────────────────────────────────────────────────────
@app.route("/admin/login", methods=["GET","POST"])
@rate_limit(max_calls=5, window=60)
def admin_login():
    if current_user.is_authenticated and current_user.is_admin:
        return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        username = sanitize(request.form.get("username",""))
        password = request.form.get("password","")
        ip, ua   = request.remote_addr, request.user_agent.string[:200]
        with get_db() as db:
            row = db.execute(
                "SELECT * FROM users WHERE username=? AND is_admin=1", (username,)
            ).fetchone()
            ok = bool(row and check_password_hash(row["password"], password))
            db.execute("""
                INSERT INTO login_history (user_id,ip,user_agent,success,created_at)
                VALUES (?,?,?,?,?)
            """, (row["id"] if row else None, ip, ua,
                  1 if ok else 0, datetime.now().isoformat()))
            if ok:
                db.execute("UPDATE users SET last_login=? WHERE id=?",
                           (datetime.now().strftime("%Y-%m-%d %H:%M"), row["id"]))
            db.commit()
        if ok:
            login_user(User(row), remember=True)
            flash(f"Admin login: {username}", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Invalid admin credentials.", "error")
    return render_template("admin_login.html")

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
@verified_required
def dashboard():
    with get_db() as db:
        subs = db.execute(
            "SELECT * FROM submissions WHERE user_id=? ORDER BY solved_date DESC",
            (current_user.id,)
        ).fetchall()

        fol_count = db.execute(
            "SELECT COUNT(*) as c FROM follows WHERE follower_id=?",
            (current_user.id,)
        ).fetchone()["c"]

        fwg_count = db.execute(
            "SELECT COUNT(*) as c FROM follows WHERE following_id=?",
            (current_user.id,)
        ).fetchone()["c"]

        sync_log = db.execute(
            "SELECT * FROM sync_logs WHERE user_id=? ORDER BY created_at DESC LIMIT 1",
            (current_user.id,)
        ).fetchone()

    # 🔥 MAIN DATA
    data = get_dashboard_data(subs)

    # 🔥 ADD COUNTS (your requirement)
    total, solved = get_counts(current_user.id)
    data["total"] = total
    data["solved"] = solved

    # 🔥 EXTRA DATA
    data["following_count"] = fol_count
    data["follower_count"] = fwg_count
    data["last_sync_msg"] = dict(sync_log) if sync_log else None

    # Daily challenges
    challenges = generate_daily_challenges(current_user.id, get_db)

    # Mentor tasks
    with get_db() as db:
        mentor_tasks = db.execute(
            "SELECT * FROM mentor_assignments WHERE user_id=? AND completed=0",
            (current_user.id,)
        ).fetchall()

    # Sheet URL
    sheet_url = None
    if current_user.sheet_id and current_user.sheet_id.strip():
        sheet_url = f"https://docs.google.com/spreadsheets/d/{current_user.sheet_id}"

    with get_db() as db:
        user = db.execute(
            "SELECT * FROM users WHERE id=?",
            (current_user.id,)
        ).fetchone()

    return render_template(
    "dashboard.html",
    user=user,
    data=data,
    challenges=challenges,
    mentor_tasks=mentor_tasks,
    sheet_url=sheet_url
)
# ── Mark challenge complete ────────────────────────────────────────────────────
@app.route("/challenge/complete/<int:cid>", methods=["POST"])
@login_required
@verified_required
def complete_challenge(cid):
    with get_db() as db:
        db.execute(
            "UPDATE daily_challenges SET completed=1 WHERE id=? AND user_id=?",
            (cid, current_user.id)
        )
        db.commit()
    return jsonify({"success": True})

@app.route("/mentor/complete/<int:mid>", methods=["POST"])
@login_required
@verified_required
def complete_mentor(mid):
    with get_db() as db:
        db.execute(
            "UPDATE mentor_assignments SET completed=1 WHERE id=? AND user_id=?",
            (mid, current_user.id)
        )
        db.commit()
    return jsonify({"success": True})

# ── Settings ──────────────────────────────────────────────────────────────────
@app.route("/settings", methods=["GET","POST"])
@login_required
@verified_required
def settings():
    if request.method == "POST":
        action = request.form.get("action","profile")
        with get_db() as db:
            if action == "profile":
                cf   = sanitize(request.form.get("cf_handle",""), 64)
                lc   = sanitize(request.form.get("lc_handle",""), 64)
                lc_pw     = request.form.get("lc_password","").strip()[:128]
                lc_cookie = request.form.get("lc_session_cookie","").strip()[:500]
                lc_csrf   = request.form.get("lc_csrf_token","").strip()[:100]
                ac   = sanitize(request.form.get("ac_handle",""), 64)
                sid  = sanitize(request.form.get("sheet_id",""), 128)
                bio  = sanitize(request.form.get("bio",""), 300)
                pub  = 1 if request.form.get("is_public") else 0
                plat = request.form.getlist("platforms")
                if not plat:
                    plat = ["Codeforces","LeetCode","AtCoder"]
                valid_plat = [p for p in plat
                              if p in ["Codeforces","LeetCode","AtCoder"]]
                db.execute("""
                    UPDATE users SET cf_handle=?,lc_handle=?,lc_password=?,
                    lc_session_cookie=?,lc_csrf_token=?,ac_handle=?,sheet_id=?,bio=?,is_public=?,
                    enabled_platforms=? WHERE id=?
                """, (cf, lc,
                      generate_password_hash(lc_pw) if lc_pw else current_user.lc_password,
                      lc_cookie, lc_csrf,
                      ac, sid, bio, pub,
                      json.dumps(valid_plat), current_user.id))
                flash("Profile settings saved!", "success")
                print("COOKIE:", lc_cookie[:20])
                print("CSRF:", lc_csrf)

            elif action == "password":
                old  = request.form.get("old_password","")
                new  = request.form.get("new_password","")
                conf = request.form.get("confirm_password","")
                row  = db.execute("SELECT password FROM users WHERE id=?",
                                  (current_user.id,)).fetchone()
                if not check_password_hash(row["password"], old):
                    flash("Current password incorrect.", "error")
                elif new != conf:
                    flash("New passwords don't match.", "error")
                elif len(new) < 6:
                    flash("Min 6 characters.", "error")
                else:
                    db.execute("UPDATE users SET password=? WHERE id=?",
                               (generate_password_hash(new), current_user.id))
                    flash("Password changed!", "success")
            db.commit()
        return redirect(url_for("settings"))
    return render_template("settings.html", user=current_user)

#---leetcode connect-------------------------------------------------------------

@app.route("/leetcode/connect", methods=["POST"])
@login_required
def leetcode_connect():

    if webdriver is None or Service is None or ChromeDriverManager is None:
        return jsonify({"success": False, "message": "Browser automation dependencies are not installed."}), 503

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install())
    )


    driver.get("https://leetcode.com/accounts/login/")

    for _ in range(120):

        cookies = driver.get_cookies()
        names = [c["name"] for c in cookies]

        if "LEETCODE_SESSION" in names:
            break

        time.sleep(1)

    if "LEETCODE_SESSION" not in names:
        driver.quit()

        return jsonify({
        "success": False,
        "message": "Login timeout. Please try again."
    })
    lc_session = None

    csrf_token = None

    for cookie in cookies:

        if cookie["name"] == "LEETCODE_SESSION":
            lc_session = cookie["value"]

        elif cookie["name"] == "csrftoken":
            csrf_token = cookie["value"]

    if lc_session and csrf_token:

        with get_db() as db:
            db.execute(
                """
                UPDATE users
                SET lc_session_cookie=?,
                    lc_csrf_token=?
                WHERE id=?
                """,
                (
                    lc_session,
                    csrf_token,
                    current_user.id
                )
            )
            db.commit()

        driver.quit()

        return jsonify({
            "success": True,
            "message": "LeetCode Connected"
        })

    driver.quit()

    return jsonify({
        "success": False,
        "message": "Login Failed"
    })
# ── Sync ──────────────────────────────────────────────────────────────────────

@app.route("/sync", methods=["POST"])
@login_required
@verified_required
def sync():
    with get_db() as db:
        user = db.execute(
            "SELECT * FROM users WHERE id=?", (current_user.id,)
        ).fetchone()

    user = dict(user)

    # ❌ block if not imported
    if int(user.get("lc_imported", 0)) == 0:
        return jsonify({
            "success": False,
            "message": "❌ Please import LeetCode first"
        })

    # 🔥 Daily sync should not re-fetch LeetCode.
    # Full import is handled by /import_lc and bot_sheet_sync.py.

    result = sync_user_data(user, get_db)
    return jsonify(result)

#________________________________________________________________________________________
def _do_import(user_id):

    print("=" * 50)
    print("🚀 _do_import STARTED")
    print("USER ID =", user_id)
    print("=" * 50)

    import subprocess
    import sys

    result = subprocess.run(
        [sys.executable, "bot_sheet_sync.py", str(user_id), "import"],
        capture_output=False
    )

    print("RETURN CODE =", result.returncode)
    
# ── LeetCode Full Import (First Time — uses LEETCODE_SESSION cookie) ──────────
@app.route("/import_lc", methods=["POST"])
@login_required
def import_lc():

    with get_db() as db:
        user = db.execute(
            "SELECT * FROM users WHERE id=?",
            (current_user.id,)
        ).fetchone()

    if not user["lc_session_cookie"]:
        return jsonify({
            "success": False,
            "message": "Connect LeetCode first"
        })

    threading.Thread(
        target=_do_import,
        args=(current_user.id,),
        daemon=True
    ).start()

    return jsonify({
        "success": True,
        "message": "Import started"
    })
# ── Problems Table ────────────────────────────────────────────────────────────
@app.route("/problems")
@login_required
@verified_required
def problems():
    platform = request.args.get("platform","all")
    diff     = request.args.get("diff","all")
    search   = sanitize(request.args.get("q",""), 64)
    page     = max(1, int(request.args.get("page",1)))
    per_page = 50

    with get_db() as db:
        base   = "SELECT * FROM submissions WHERE user_id=?"
        params = [current_user.id]
        if platform != "all":
            base += " AND platform=?"; params.append(platform)
        if diff != "all":
            base += " AND difficulty=?"; params.append(diff)
        if search:
            base += " AND (problem_name LIKE ? OR tags LIKE ?)"; params += [f"%{search}%", f"%{search}%"]
        total = db.execute(f"SELECT COUNT(*) as c FROM ({base})", params).fetchone()["c"]
        base += " ORDER BY solved_date DESC LIMIT ? OFFSET ?"
        params += [per_page, (page-1)*per_page]
        subs   = db.execute(base, params).fetchall()

    return render_template("problems.html",
                           subs=subs, total=total, page=page,
                           total_pages=(total+per_page-1)//per_page,
                           platform=platform, diff=diff, search=search)

# ── Export XLSX ───────────────────────────────────────────────────────────────
@app.route("/export_csv")
@login_required
@verified_required
def export_csv():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    with get_db() as db:
        subs = db.execute(
            "SELECT solved_date,problem_name,problem_url,difficulty,platform,tags "
            "FROM submissions WHERE user_id=? ORDER BY solved_date DESC",
            (current_user.id,)
        ).fetchall()
    topic_counter: Counter = Counter()
    for s in subs:
        tags = s["tags"] or ""
        first = next((t.strip() for t in tags.split(";") if t.strip()), "General")
        topic_counter[first] += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{current_user.username} DSA"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    headers = ["DATE", "PROGRAM TITLE", "LINK", "DIFFICULTY", "PLATFORM", "TOPIC", "COUNT"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    diff_colors = {"Easy": "C6EFCE", "Medium": "FFEB9C", "Hard": "FFC7CE"}
    for s in subs:
        tags  = s["tags"] or ""
        first = next((t.strip() for t in tags.split(";") if t.strip()), "General")
        name  = str(s["problem_name"] or "").lstrip("=+-@")
        ws.append([s["solved_date"], name, s["problem_url"], s["difficulty"],
                   s["platform"], first, topic_counter[first]])
        diff = s["difficulty"] or ""
        color = diff_colors.get(diff)
        if color:
            ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=color)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"dsa_{current_user.username}.xlsx")

# ── Public Profile ────────────────────────────────────────────────────────────
@app.route("/user/<username>")
def public_profile(username):
    with get_db() as db:
        row = db.execute(
            "SELECT * FROM users WHERE username=? AND status='active'", (username,)
        ).fetchone()
        if not row: abort(404)
        user = User(row)
        self_view  = current_user.is_authenticated and current_user.id == user.id
        admin_view = current_user.is_authenticated and current_user.is_admin
        if not user.is_public and not self_view and not admin_view:
            abort(403)
        subs = db.execute("""
SELECT * FROM submissions 
WHERE user_id=? 
ORDER BY 
    substr(solved_date,7,4) DESC,
    substr(solved_date,4,2) DESC,
    substr(solved_date,1,2) DESC
""", (current_user.id,)).fetchall()
        fol_c = db.execute("SELECT COUNT(*) as c FROM follows WHERE following_id=?",
                           (user.id,)).fetchone()["c"]
        fwg_c = db.execute("SELECT COUNT(*) as c FROM follows WHERE follower_id=?",
                           (user.id,)).fetchone()["c"]
        is_following = False
        if current_user.is_authenticated and not self_view:
            is_following = bool(db.execute(
                "SELECT 1 FROM follows WHERE follower_id=? AND following_id=?",
                (current_user.id, user.id)
            ).fetchone())
    data = get_dashboard_data(subs)
    total, solved = get_counts(current_user.id)
    data["total"] = total
    data["solved"] = solved
    return render_template("profile.html",
                           profile_user=user, data=data, subs=list(subs)[:20],
                           follower_count=fol_c, following_count=fwg_c,
                           is_following=is_following, viewer_is_self=self_view)

# ── Follow ─────────────────────────────────────────────────────────────────────
@app.route("/follow/<int:uid>", methods=["POST"])
@login_required
@verified_required
def follow(uid):
    if uid == current_user.id:
        return jsonify({"success": False, "message": "Cannot follow yourself."})
    with get_db() as db:
        ex = db.execute("SELECT 1 FROM follows WHERE follower_id=? AND following_id=?",
                        (current_user.id, uid)).fetchone()
        if ex:
            db.execute("DELETE FROM follows WHERE follower_id=? AND following_id=?",
                       (current_user.id, uid))
            action = "unfollowed"
        else:
            db.execute("INSERT INTO follows (follower_id,following_id,created_at) VALUES (?,?,?)",
                       (current_user.id, uid, datetime.now().isoformat()))
            action = "followed"
        cnt = db.execute("SELECT COUNT(*) as c FROM follows WHERE following_id=?",
                         (uid,)).fetchone()["c"]
        db.commit()
    return jsonify({"success": True, "action": action, "count": cnt})

# ── Friends / Leaderboard ──────────────────────────────────────────────────────
@app.route("/friends")
@login_required
@verified_required
def friends():
    with get_db() as db:
        following = db.execute("""
            SELECT u.id,u.username,u.cf_handle,u.lc_handle,u.ac_handle,
                   u.last_sync,u.is_public,
                   (SELECT COUNT(*) FROM submissions WHERE user_id=u.id) as prob_count
            FROM follows f JOIN users u ON f.following_id=u.id
            WHERE f.follower_id=? AND u.status='active'
            ORDER BY prob_count DESC
        """, (current_user.id,)).fetchall()
        followers = db.execute("""
            SELECT u.id,u.username,u.is_public,
                   (SELECT COUNT(*) FROM submissions WHERE user_id=u.id) as prob_count
            FROM follows f JOIN users u ON f.follower_id=u.id
            WHERE f.following_id=? AND u.status='active'
        """, (current_user.id,)).fetchall()
    return render_template("friends.html", following=following, followers=followers)

@app.route("/leaderboard")
@login_required
@verified_required
def leaderboard():
    with get_db() as db:
        users = db.execute("""
            SELECT u.id,u.username,u.is_public,
                   COUNT(s.id) as total,
                   SUM(CASE WHEN s.platform='Codeforces' THEN 1 ELSE 0 END) as cf,
                   SUM(CASE WHEN s.platform='LeetCode'   THEN 1 ELSE 0 END) as lc,
                   SUM(CASE WHEN s.platform='AtCoder'    THEN 1 ELSE 0 END) as ac,
                   u.last_sync
            FROM users u
            LEFT JOIN submissions s ON u.id=s.user_id
            WHERE u.status='active' AND u.is_verified=1 AND u.is_admin=0
            GROUP BY u.id ORDER BY total DESC LIMIT 50
        """).fetchall()
    return render_template("leaderboard.html", users=users)

@app.route("/login_history")
@login_required
@verified_required
def login_history():
    with get_db() as db:
        history = db.execute(
            "SELECT * FROM login_history WHERE user_id=? ORDER BY created_at DESC LIMIT 30",
            (current_user.id,)
        ).fetchall()
    return render_template("login_history.html", history=history)

# ── Admin Dashboard ────────────────────────────────────────────────────────────
@app.route("/admin")
@login_required
@admin_required
def admin_dashboard():


    with get_db() as db:
        users   = db.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
        counts  = {u["id"]: db.execute(
            "SELECT COUNT(*) as c FROM submissions WHERE user_id=?",
            (u["id"],)
        ).fetchone()["c"] for u in users}
        pending = db.execute(
            "SELECT COUNT(*) as c FROM users WHERE status='pending'"
        ).fetchone()["c"]
        total_subs = db.execute(
            "SELECT COUNT(*) as c FROM submissions"
        ).fetchone()["c"]
        logs = db.execute(
            "SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT 30"
        ).fetchall()

        # Analytics: signups per day (last 14 days)
        signups_raw = db.execute("""
            SELECT DATE(created_at) as day, COUNT(*) as cnt
            FROM users WHERE created_at >= DATE('now','-14 days')
            GROUP BY day ORDER BY day
        """).fetchall()
        # Problems solved per day (last 14 days)
        probs_raw = db.execute("""
            SELECT solved_date, COUNT(*) as cnt
            FROM submissions
            GROUP BY solved_date
            ORDER BY solved_date DESC LIMIT 14
        """).fetchall()

        active_users = db.execute(
            "SELECT COUNT(*) as c FROM users WHERE status='active'"
        ).fetchone()["c"]

        custom_probs = db.execute(
            "SELECT * FROM custom_problems ORDER BY created_at DESC"
        ).fetchall()

    signup_labels = [r["day"] for r in signups_raw]
    signup_data   = [r["cnt"] for r in signups_raw]
    prob_labels   = [r["solved_date"] for r in probs_raw]
    prob_data     = [r["cnt"] for r in probs_raw]

    return render_template("admin.html",
                           users=users, counts=counts, pending=pending,
                           total_subs=total_subs, logs=logs,
                           active_users=active_users,
                           signup_labels=signup_labels, signup_data=signup_data,
                           prob_labels=prob_labels, prob_data=prob_data,
                           custom_probs=custom_probs)

@app.route("/admin/verify/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_verify(uid):
    action = request.form.get("action","approve")
    with get_db() as db:
        u = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
        if not u: return redirect(url_for("admin_dashboard"))
        if action == "approve":
            db.execute("UPDATE users SET is_verified=1,status='active' WHERE id=?", (uid,))
        else:
            db.execute("UPDATE users SET status='rejected' WHERE id=?", (uid,))
        db.commit()
    log_admin("verify_user", u["username"], action)
    flash(f"User {u['username']} {action}d.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/delete/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_delete(uid):
    if uid == current_user.id:
        flash("Cannot delete yourself.", "error")
        return redirect(url_for("admin_dashboard"))
    with get_db() as db:
        u = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
        if u:
            for tbl in ["submissions","follows","login_history","sync_logs",
                        "daily_challenges","mentor_assignments"]:
                db.execute(f"DELETE FROM {tbl} WHERE user_id=?", (uid,))
            db.execute("DELETE FROM follows WHERE following_id=?", (uid,))
            db.execute("DELETE FROM users WHERE id=?", (uid,))
            db.commit()
            log_admin("delete_user", u["username"])
    flash("User deleted.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/promote/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_promote(uid):
    with get_db() as db:
        row = db.execute("SELECT username,is_admin FROM users WHERE id=?", (uid,)).fetchone()
        if row:
            nv = 0 if row["is_admin"] else 1
            db.execute("UPDATE users SET is_admin=? WHERE id=?", (nv, uid))
            db.commit()
            log_admin("promote_user", row["username"],
                      "to admin" if nv else "to user")
    flash("Role updated.", "success")
    return redirect(url_for("admin_dashboard"))

# ── Mentor Mode ───────────────────────────────────────────────────────────────
@app.route("/admin/mentor", methods=["GET","POST"])
@login_required
@admin_required
def mentor():
    if request.method == "POST":
        uid   = int(request.form.get("user_id",0))
        pname = sanitize(request.form.get("problem_name",""), 200)
        purl  = sanitize(request.form.get("problem_url",""), 300)
        diff  = request.form.get("difficulty","Medium")
        plat  = request.form.get("platform","LeetCode")
        topic = sanitize(request.form.get("topic",""), 64)
        note  = sanitize(request.form.get("note",""), 500)
        due   = request.form.get("due_date","")
        with get_db() as db:
            db.execute("""
                INSERT INTO mentor_assignments
                (admin_id,user_id,problem_name,problem_url,difficulty,
                 platform,topic,note,assigned_date,due_date,completed)
                VALUES (?,?,?,?,?,?,?,?,?,?,0)
            """, (current_user.id, uid, pname, purl, diff, plat,
                  topic, note, datetime.now().strftime("%Y-%m-%d"), due))
            db.commit()
        flash(f"Problem assigned.", "success")
        return redirect(url_for("mentor"))

    with get_db() as db:
        users = db.execute(
            "SELECT id,username FROM users WHERE status='active' AND is_admin=0"
        ).fetchall()
        assignments = db.execute("""
            SELECT ma.*,u.username
            FROM mentor_assignments ma
            JOIN users u ON ma.user_id=u.id
            ORDER BY ma.assigned_date DESC LIMIT 50
        """).fetchall()
        custom_probs = db.execute(
            "SELECT * FROM custom_problems ORDER BY created_at DESC"
        ).fetchall()

        # Build user_progress: total assigned vs completed per user
        user_progress = {}
        for u in users:
            uid = u["id"]
            total_assigned = db.execute(
                "SELECT COUNT(*) as c FROM mentor_assignments WHERE user_id=?", (uid,)
            ).fetchone()["c"]
            total_completed = db.execute(
                "SELECT COUNT(*) as c FROM mentor_assignments WHERE user_id=? AND completed=1", (uid,)
            ).fetchone()["c"]
            pct = int((total_completed / total_assigned * 100)) if total_assigned else 0
            user_progress[uid] = {
                "total": total_assigned,
                "completed": total_completed,
                "pct": pct
            }

    return render_template("mentor.html", users=users,
                           assignments=assignments, custom_probs=custom_probs,
                           user_progress=user_progress)

# ── Custom Problems ───────────────────────────────────────────────────────────
@app.route("/admin/custom_problem", methods=["POST"])
@login_required
@admin_required
def create_custom_problem():
    title = sanitize(request.form.get("title",""), 200)
    desc  = sanitize(request.form.get("description",""), 2000)
    diff  = request.form.get("difficulty","Medium")
    topic = sanitize(request.form.get("topic",""), 64)
    url   = sanitize(request.form.get("url",""), 300)
    with get_db() as db:
        db.execute("""
            INSERT INTO custom_problems
            (created_by,title,description,difficulty,topic,url,created_at)
            VALUES (?,?,?,?,?,?,?)
        """, (current_user.id, title, desc, diff, topic, url,
              datetime.now().isoformat()))
        db.commit()
    log_admin("create_custom_problem", details=title)
    flash(f"Custom problem '{title}' created.", "success")
    return redirect(url_for("mentor"))

# ── Admin Sync All ────────────────────────────────────────────────────────────
@app.route("/admin/sync_user/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_sync_user(uid):
    def _do():
        with get_db() as db:
            u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not u:
            return
        try:
            res = sync_user_data(dict(u), get_db)
            with get_db() as db:
                db.execute("""
                    INSERT INTO sync_logs (user_id,status,message,created_at)
                    VALUES (?,?,?,?)
                """, (uid,
                      "success" if res["success"] else "error",
                      res["message"], datetime.now().isoformat()))
                if res["success"]:
                    db.execute("UPDATE users SET last_sync=? WHERE id=?",
                               (datetime.now().strftime("%Y-%m-%d %H:%M"), uid))
                db.commit()
        except Exception as e:
            print(f"[SyncUser] {e}")
    threading.Thread(target=_do, daemon=True).start()
    log_admin("sync_user", details=f"uid={uid}")
    return jsonify({"ok": True, "message": "Sync started"})

@app.route("/admin/reset_lc/<int:uid>", methods=["POST"])
@login_required
@admin_required
def admin_reset_lc(uid):
    with get_db() as db:
        u = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
        if u:
            db.execute("UPDATE users SET lc_imported=0 WHERE id=?", (uid,))
            db.commit()
            log_admin("reset_lc_import", u["username"])
    return jsonify({"ok": True})

@app.route("/admin/sync_all", methods=["POST"])
@login_required
@admin_required
def admin_sync_all():
    def _do():
        with get_db() as db:
            users = db.execute(
                "SELECT * FROM users WHERE status='active' AND is_verified=1"
            ).fetchall()
        for u in users:
            try:
                res = sync_user_data(u, get_db)
                with get_db() as db:
                    db.execute("""
                        INSERT INTO sync_logs (user_id,status,message,created_at)
                        VALUES (?,?,?,?)
                    """, (u["id"],
                          "success" if res["success"] else "error",
                          res["message"], datetime.now().isoformat()))
                    if res["success"]:
                        db.execute("UPDATE users SET last_sync=? WHERE id=?",
                                   (datetime.now().strftime("%Y-%m-%d %H:%M"), u["id"]))
                    db.commit()
                time.sleep(3)
            except Exception as e:
                print(f"[AutoSync] {u['username']}: {e}")
    threading.Thread(target=_do, daemon=True).start()
    log_admin("sync_all", details="Background sync triggered")
    flash("Background sync started for all users.", "success")
    return redirect(url_for("admin_dashboard"))

# ── Backup / Restore ──────────────────────────────────────────────────────────
@app.route("/admin/backup", methods=["POST"])
@login_required
@admin_required
def admin_backup():
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"backup_{ts}.db")
    shutil.copy2(DB, dst)
    log_admin("backup", details=dst)
    flash(f"Backup created: backup_{ts}.db", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/restore", methods=["POST"])
@login_required
@admin_required
def admin_restore():
    fname = request.form.get("backup_file","")
    if not fname or ".." in fname or "/" in fname:
        flash("Invalid file.", "error")
        return redirect(url_for("admin_dashboard"))
    src = os.path.join(BACKUP_DIR, fname)
    if not os.path.exists(src):
        flash("File not found.", "error")
        return redirect(url_for("admin_dashboard"))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(DB, os.path.join(BACKUP_DIR, f"pre_restore_{ts}.db"))
    shutil.copy2(src, DB)
    log_admin("restore", details=fname)
    flash(f"Restored from {fname}.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/backups_list")
@login_required
@admin_required
def admin_backups_list():
    files = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")], reverse=True)
    return jsonify(files)

# ── Auto-sync background thread (stable single instance) ─────────────────────

# ── Error handlers ────────────────────────────────────────────────────────────
@app.errorhandler(404)
def e404(e): return render_template("error.html", code=404, msg="Page not found."), 404
@app.errorhandler(403)
def e403(e): return render_template("error.html", code=403, msg="Access denied."), 403
@app.errorhandler(429)
def e429(e): return render_template("error.html", code=429, msg="Too many requests."), 429


def ensure_db_columns():
    with get_db() as db:
        for stmt in [
            "ALTER TABLE users ADD COLUMN lc_imported INTEGER DEFAULT 0",
            "ALTER TABLE users ADD COLUMN auto_sync_enabled INTEGER DEFAULT 1",
            "ALTER TABLE users ADD COLUMN sync_time TEXT DEFAULT '09:00'",
        ]:
            try:
                db.execute(stmt)
            except:
                pass
        db.commit()

#--------------------------------------------------------------------------

import threading

def schedule_user(u):
    while True:
        now = datetime.now()

        sync_time = u.get("sync_time")
        if not sync_time:
            return

        target_time = datetime.strptime(sync_time, "%H:%M").replace(
            year=now.year,
            month=now.month,
            day=now.day
        )

        if target_time <= now:
            target_time += timedelta(days=1)

        wait_seconds = (target_time - now).total_seconds()

        print(f"⏳ {u['username']} waiting {int(wait_seconds)} sec")

        time.sleep(wait_seconds)

        print(f"⏰ Running auto sync for {u['username']}")
        sync_user_data(u, get_db)
        time.sleep(60)


def run_scheduler():
    while True:
        now = datetime.now().strftime("%H:%M")

        with get_db() as db:
            users = db.execute(
                "SELECT * FROM users WHERE lc_imported=1 AND COALESCE(auto_sync_enabled,1)=1"
            ).fetchall()

        for u in users:
            u = dict(u)

            last_run = u.get("last_sync", "")

            if last_run:
                last_run = last_run[-5:]  # HH:MM only

            if u.get("sync_time") == now and last_run != now:
                print(f"⏰ Running auto sync for {u['username']}")

                sync_user_data(u, get_db)

                with get_db() as db:
                    db.execute(
                        "UPDATE users SET last_sync=? WHERE id=?",
                        (datetime.now().strftime("%Y-%m-%d %H:%M"), u["id"])
                    )
                    db.commit()

        time.sleep(60)  # 🔥 THIS MUST ALIGN WITH while


#-----------------------------------------------------------------------
@app.route("/set_sync_time", methods=["POST"])
@login_required
def set_sync_time():
    data = request.json or {}
    sync_time = data.get("time", "09:00")

    with get_db() as db:
        db.execute(
            "UPDATE users SET sync_time=? WHERE id=?",
            (sync_time, current_user.id)
        )
        db.commit()

    return jsonify({"success": True, "sync_time": sync_time})


#------[Feed Back Section] --------------------------------------------------
import os
import smtplib
from email.mime.text import MIMEText
from flask import request, jsonify
from flask_login import login_required, current_user


@app.route("/feedback", methods=["POST"])
@login_required
def feedback():

    try:
        data = request.get_json() or {}

        feedback_text = data.get("feedback", "").strip()

        if not feedback_text:
            return jsonify({
                "success": False,
                "message": "Feedback is empty"
            })

        sender_email = os.getenv("SMTP_EMAIL")
        sender_password = os.getenv("SMTP_APP_PASSWORD")
        owner_email = os.getenv("OWNER_EMAIL")

        msg = MIMEText(f"""
User: {current_user.username}
Email: {getattr(current_user, 'email', '')}

Feedback:

{feedback_text}
""")

        msg["Subject"] = "DSA Tracker Feedback"
        msg["From"] = sender_email
        msg["To"] = owner_email

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()

        server.login(
            sender_email,
            sender_password
        )

        server.send_message(msg)
        server.quit()

        return jsonify({
            "success": True,
            "message": "Feedback sent successfully"
        })

    except Exception as e:

        print("FEEDBACK ERROR =", e)

        return jsonify({
            "success": False,
            "message": str(e)
        })
# ---------------------------------------------------------------------------
@app.route("/toggle_auto_sync", methods=["POST"])
@login_required
def toggle_auto_sync():
    data = request.json or {}
    enabled = 1 if data.get("enabled") else 0
    with get_db() as db:
        db.execute(
            "UPDATE users SET auto_sync_enabled=? WHERE id=?",
            (enabled, current_user.id)
        )
        db.commit()
    return jsonify({"success": True, "enabled": bool(enabled)})



# ======================================================================
# Daily Tracker / Student Sheets
# Master sheet is derived from submissions (problem logs) already stored
# in the DB. Admin-created sheets are stored separately and can be auto
# evaluated or manually updated.
# ======================================================================

def _parse_any_date(value):
    """Parse dates stored in multiple formats used by the app."""
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    fmts = ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%y")
    for fmt in fmts:
        try:
            return datetime.strptime(s[:10], fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None

def _date_label(dt_obj, fallback="Unknown"):
    if not dt_obj:
        return fallback
    return dt_obj.strftime("%d-%m-%Y")

def _clean_topic(topic):
    return " ".join(str(topic or "").split()).strip()

def _group_rows(rows, include_username=False):
    """Return grouped daily rows from submission rows."""
    groups = {}
    for row in rows:
        dt = _parse_any_date(row["solved_date"])
        date_label = _date_label(dt, str(row["solved_date"] or "Unknown"))

        username = row["username"] if include_username and "username" in row.keys() else current_user.username
        key = (date_label, username)

        if key not in groups:
            groups[key] = {
                "date": date_label,
                "date_dt": dt or datetime.min,
                "username": username,
                "solved_count": 0,
                "easy": 0,
                "medium": 0,
                "hard": 0,
                "problems": []
            }

        item = groups[key]
        item["solved_count"] += 1
        diff = str(row["difficulty"] or "").strip().lower()
        if diff == "easy":
            item["easy"] += 1
        elif diff == "medium":
            item["medium"] += 1
        elif diff == "hard":
            item["hard"] += 1

        item["problems"].append({
            "name": row["problem_name"] or "",
            "url": row["problem_url"] or "",
            "difficulty": row["difficulty"] or "",
            "platform": row["platform"] or "",
            "topic": row["tags"] or row["topic"] if "topic" in row.keys() else (row["tags"] or "")
        })

    # Sort: latest date first, then count desc, then username
    ordered = sorted(
        groups.values(),
        key=lambda x: (
            x["date_dt"] if x["date_dt"] != datetime.min else datetime.min,
            x["solved_count"],
            x["username"].lower()
        ),
        reverse=True
    )
    return ordered

def _get_user_by_username(username):
    with get_db() as db:
        row = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    return row

def _tracker_defaults():
    return [
        ("Daily Problems", "daily_problems", "Auto-filled from submissions / problem logs", 0, ""),
        ("Weekly Assignment", "weekly_assignment", "Auto-evaluated from the last 7 days", 10, ""),
        ("Contest Attendance", "contest_attendance", "Manual / admin verified contest sheet", 0, ""),
        ("Monthly Goal", "monthly_goal", "Auto-evaluated using current month submissions", 50, ""),
        ("Placement Eligibility", "placement_eligibility", "Auto-evaluated from total solved & hard problems", 300, "50"),
        ("Topic Verification", "topic_verification", "Auto-evaluated by topic count", 3, "Graphs"),
        ("Custom Sheets", "custom", "Admin-defined custom verification sheet", 0, "")
    ]

def ensure_tracker_schema():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS daily_tracker_sheets (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT UNIQUE NOT NULL,
            sheet_type    TEXT NOT NULL DEFAULT 'custom',
            description   TEXT DEFAULT '',
            target_value  INTEGER DEFAULT 0,
            target_topic  TEXT DEFAULT '',
            created_by    INTEGER DEFAULT 0,
            visibility    TEXT DEFAULT 'private',
            created_at    TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS daily_tracker_sheet_results (
            sheet_id     INTEGER NOT NULL,
            student_id   INTEGER NOT NULL,
            status       TEXT DEFAULT '',
            solved_count INTEGER DEFAULT 0,
            details_json  TEXT DEFAULT '',
            remarks      TEXT DEFAULT '',
            updated_at   TEXT DEFAULT '',
            PRIMARY KEY (sheet_id, student_id)
        );
        """)
        # Seed default tracker sheets once
        for name, sheet_type, description, target_value, target_topic in _tracker_defaults():
            exists = db.execute(
                "SELECT 1 FROM daily_tracker_sheets WHERE name=?",
                (name,)
            ).fetchone()
            if not exists:
                db.execute("""
                    INSERT INTO daily_tracker_sheets
                    (name, sheet_type, description, target_value, target_topic, created_by, visibility, created_at)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (
                    name, sheet_type, description, int(target_value or 0), str(target_topic or ""),
                    0, "private", datetime.now().isoformat(timespec="seconds")
                ))
        db.commit()

def _tracker_metrics_for_users():
    """
    Build per-user metrics from submissions so the tracker can compute
    weekly/monthly/topic/placement sheets.
    """
    today = datetime.now().date()
    week_start = today - timedelta(days=6)
    month_start = today.replace(day=1)

    with get_db() as db:
        rows = db.execute("""
            SELECT s.user_id, s.platform, s.problem_name, s.problem_id, s.problem_url,
                   s.difficulty, s.tags, s.solved_date, u.username
            FROM submissions s
            JOIN users u ON u.id = s.user_id
            WHERE u.status='active' AND u.is_admin=0
            ORDER BY s.solved_date DESC, s.id DESC
        """).fetchall()

    metrics = {}
    for row in rows:
        uid = row["user_id"]
        if uid not in metrics:
            metrics[uid] = {
                "username": row["username"],
                "total": 0,
                "today": 0,
                "week": 0,
                "month": 0,
                "hard": 0,
                "topics": {},
                "rows": []
            }

        dt = _parse_any_date(row["solved_date"])
        if not dt:
            continue
        d = dt.date()
        topic = _clean_topic(row["tags"])
        difficulty = (row["difficulty"] or "").strip().lower()

        metrics[uid]["total"] += 1
        if d == today:
            metrics[uid]["today"] += 1
        if d >= week_start:
            metrics[uid]["week"] += 1
        if d >= month_start:
            metrics[uid]["month"] += 1
        if difficulty == "hard":
            metrics[uid]["hard"] += 1

        if topic:
            metrics[uid]["topics"][topic] = metrics[uid]["topics"].get(topic, 0) + 1

        metrics[uid]["rows"].append(row)

    return metrics

def refresh_tracker_sheet_results():
    """
    Auto-compute results for the built-in verification sheets.
    Manual contest/custom sheets can still be edited by admin.
    """
    metrics = _tracker_metrics_for_users()
    today = datetime.now().date()

    with get_db() as db:
        sheets = db.execute("SELECT * FROM daily_tracker_sheets ORDER BY id ASC").fetchall()
        active_users = db.execute("""
            SELECT id, username FROM users
            WHERE status='active' AND is_admin=0
            ORDER BY username ASC
        """).fetchall()

        for sheet in sheets:
            stype = (sheet["sheet_type"] or "").strip().lower()
            if stype not in {"weekly_assignment", "monthly_goal", "placement_eligibility", "topic_verification", "coding_attendance"}:
                continue

            target_value = int(sheet["target_value"] or 0)
            target_topic = str(sheet["target_topic"] or "").strip()

            for user in active_users:
                m = metrics.get(user["id"], {
                    "total": 0, "today": 0, "week": 0, "month": 0, "hard": 0, "topics": {}
                })

                status = "NO"
                solved_count = 0
                remarks = ""

                if stype == "weekly_assignment":
                    solved_count = int(m["week"])
                    status = "YES" if solved_count >= target_value else "NO"
                    remarks = f"Weekly solved: {solved_count}/{target_value}"
                elif stype == "monthly_goal":
                    solved_count = int(m["month"])
                    status = "YES" if solved_count >= target_value else "NO"
                    remarks = f"Monthly solved: {solved_count}/{target_value}"
                elif stype == "placement_eligibility":
                    solved_count = int(m["total"])
                    hard_target = int(target_topic) if str(target_topic).isdigit() else 50
                    hard_ok = int(m["hard"]) >= hard_target
                    solved_ok = solved_count >= (target_value or 300)
                    status = "YES" if (solved_ok and hard_ok) else "NO"
                    remarks = f"Total: {solved_count}, Hard: {m['hard']} (need {hard_target})"
                elif stype == "topic_verification":
                    solved_count = int(m["topics"].get(target_topic, 0))
                    status = "YES" if solved_count >= target_value else "NO"
                    remarks = f"{target_topic}: {solved_count}/{target_value}"
                elif stype == "coding_attendance":
                    solved_count = int(m["today"])
                    status = "Present" if solved_count >= 1 else "Absent"
                    remarks = f"Today's solved count: {solved_count}"

                db.execute("""
                    INSERT INTO daily_tracker_sheet_results
                    (sheet_id, student_id, status, solved_count, details_json, remarks, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(sheet_id, student_id) DO UPDATE SET
                        status=excluded.status,
                        solved_count=excluded.solved_count,
                        details_json=excluded.details_json,
                        remarks=excluded.remarks,
                        updated_at=excluded.updated_at
                """, (
                    sheet["id"], user["id"], status, solved_count,
                    json.dumps({
                        "sheet_type": stype,
                        "target_value": target_value,
                        "target_topic": target_topic
                    }),
                    remarks,
                    datetime.now().isoformat(timespec="seconds")
                ))

        db.commit()

@app.route("/daily_tracker")
@login_required
@verified_required
def daily_tracker():
    """
    Master daily tracker page.
    Admin sees all students; regular users are redirected to their own sheet.
    """
    if not current_user.is_admin:
        return redirect(url_for("my_sheet"))

    ensure_tracker_schema()
    refresh_tracker_sheet_results()

    with get_db() as db:
        users = db.execute("""
            SELECT id, username FROM users
            WHERE status='active' AND is_admin=0
            ORDER BY username ASC
        """).fetchall()
        sheets = db.execute("""
            SELECT * FROM daily_tracker_sheets
            ORDER BY id ASC
        """).fetchall()
        results = db.execute("""
            SELECT r.*, s.name as sheet_name, u.username
            FROM daily_tracker_sheet_results r
            JOIN daily_tracker_sheets s ON s.id=r.sheet_id
            JOIN users u ON u.id=r.student_id
            ORDER BY r.updated_at DESC, u.username ASC
        """).fetchall()
        subs = db.execute("""
            SELECT s.*, u.username
            FROM submissions s
            JOIN users u ON u.id=s.user_id
            WHERE u.status='active' AND u.is_admin=0
            ORDER BY s.solved_date DESC, u.username ASC, s.id DESC
        """).fetchall()

    daily_rows = _group_rows(subs, include_username=True)

    today_label_iso = datetime.now().strftime("%Y-%m-%d")
    today_label_dmy = datetime.now().strftime("%d-%m-%Y")  # legacy compat
    active_today = len({row["username"] for row in daily_rows
                        if row["date"] in (today_label_iso, today_label_dmy)})

    summary_cards = {
        "students": len(users),
        "active_today": active_today,
        "total_solved": len(subs),
        "sheets": len(sheets)
    }

    return render_template(
        "daily_tracker.html",
        is_admin=True,
        view_username=None,
        sheets=sheets,
        results=results,
        daily_rows=daily_rows,
        summary_cards=summary_cards,
        users=users,
        page_title="Master Daily Tracker"
    )

@app.route("/my_sheet")
@login_required
@verified_required
def my_sheet():
    return redirect(url_for("student_sheet", username=current_user.username))

@app.route("/student_sheet/<username>")
@login_required
@verified_required
def student_sheet(username):
    ensure_tracker_schema()
    target = _get_user_by_username(username)
    if not target:
        abort(404)

    self_view = current_user.id == target["id"]
    admin_view = bool(current_user.is_admin)

    if not self_view and not admin_view:
        abort(403)

    refresh_tracker_sheet_results()

    with get_db() as db:
        subs = db.execute("""
            SELECT * FROM submissions
            WHERE user_id=?
            ORDER BY solved_date DESC, id DESC
        """, (target["id"],)).fetchall()

        sheets = db.execute("""
            SELECT * FROM daily_tracker_sheets
            ORDER BY id ASC
        """).fetchall()

        results = db.execute("""
            SELECT r.*, s.name as sheet_name
            FROM daily_tracker_sheet_results r
            JOIN daily_tracker_sheets s ON s.id=r.sheet_id
            WHERE r.student_id=?
            ORDER BY s.id ASC
        """, (target["id"],)).fetchall()

    daily_rows = _group_rows(subs, include_username=False)

    today_str = datetime.now().strftime("%Y-%m-%d")
    today_subs = [s for s in subs if (s["solved_date"] or "").startswith(today_str)]
    last_sub = subs[0] if subs else None

    # Find target for Daily Problems sheet
    daily_sheet = next((s for s in sheets if s["sheet_type"] == "daily_problems"), None)
    daily_target = int(daily_sheet["target_value"]) if daily_sheet and daily_sheet["target_value"] else 10

    summary_cards = {
        "total_solved": len(subs),
        "active_days": len(daily_rows),
        "current_streak": 0,
        "longest_streak": 0,
        "sheets": len(sheets),
        "today_solved": len(today_subs),
        "today_target": daily_target,
        "today_remaining": max(0, daily_target - len(today_subs)),
        "last_problem": last_sub["problem_name"] if last_sub else None,
        "last_solved_date": last_sub["solved_date"] if last_sub else None,
    }

    # simple streak calc from submission dates
    parsed_dates = sorted({(_parse_any_date(r["solved_date"]).date()) for r in subs if _parse_any_date(r["solved_date"])})
    if parsed_dates:
        # current streak = consecutive days ending with the most recent active day
        current = 1
        for prev, cur in zip(parsed_dates[-2::-1], parsed_dates[::-1]):
            if (cur - prev).days == 1:
                current += 1
            else:
                break

        longest = 1
        run = 1
        for prev, cur in zip(parsed_dates[:-1], parsed_dates[1:]):
            if (cur - prev).days == 1:
                run += 1
                longest = max(longest, run)
            else:
                run = 1

        summary_cards["current_streak"] = current
        summary_cards["longest_streak"] = longest

    return render_template(
        "daily_tracker.html",
        is_admin=False,
        view_username=target["username"],
        target_user=target,
        sheets=sheets,
        results=results,
        daily_rows=daily_rows,
        summary_cards=summary_cards,
        users=[],
        page_title=f"{target['username']} – My Sheet",
        self_view=self_view,
        admin_view=admin_view
    )

@app.route("/admin/tracker_sheets/create", methods=["POST"])
@login_required
@admin_required
def create_tracker_sheet():
    ensure_tracker_schema()
    name = sanitize(request.form.get("name", ""), 80)
    sheet_type = sanitize(request.form.get("sheet_type", "custom"), 40).lower()
    description = sanitize(request.form.get("description", ""), 200)
    target_value = request.form.get("target_value", "0").strip()
    target_topic = sanitize(request.form.get("target_topic", ""), 80)
    visibility = sanitize(request.form.get("visibility", "private"), 20).lower()

    if not name:
        flash("Sheet name is required.", "error")
        return redirect(url_for("daily_tracker"))

    try:
        target_value = int(target_value or 0)
    except Exception:
        target_value = 0

    with get_db() as db:
        try:
            db.execute("""
                INSERT INTO daily_tracker_sheets
                (name, sheet_type, description, target_value, target_topic, created_by, visibility, created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (
                name, sheet_type, description, target_value, target_topic,
                current_user.id, visibility, datetime.now().isoformat(timespec="seconds")
            ))
            db.commit()
            flash(f"Created sheet '{name}'.", "success")
        except Exception as e:
            flash(f"Could not create sheet: {e}", "error")

    return redirect(url_for("daily_tracker"))

@app.route("/admin/tracker_results/create", methods=["POST"])
@login_required
@admin_required
def create_tracker_result():
    ensure_tracker_schema()
    sheet_id = request.form.get("sheet_id", "").strip()
    username = sanitize(request.form.get("username", ""), 64)
    status = sanitize(request.form.get("status", ""), 40)
    remarks = sanitize(request.form.get("remarks", ""), 240)
    solved_count = request.form.get("solved_count", "0").strip()

    try:
        sheet_id = int(sheet_id)
        solved_count = int(solved_count or 0)
    except Exception:
        flash("Invalid sheet or solved count.", "error")
        return redirect(url_for("daily_tracker"))

    user = _get_user_by_username(username)
    if not user:
        flash("Username not found.", "error")
        return redirect(url_for("daily_tracker"))

    with get_db() as db:
        db.execute("""
            INSERT INTO daily_tracker_sheet_results
            (sheet_id, student_id, status, solved_count, details_json, remarks, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet_id, student_id) DO UPDATE SET
                status=excluded.status,
                solved_count=excluded.solved_count,
                details_json=excluded.details_json,
                remarks=excluded.remarks,
                updated_at=excluded.updated_at
        """, (
            sheet_id, user["id"], status, solved_count,
            json.dumps({"manual": True}),
            remarks,
            datetime.now().isoformat(timespec="seconds")
        ))
        db.commit()
    flash(f"Updated result for {username}.", "success")
    return redirect(url_for("daily_tracker"))

@app.route("/admin/tracker_refresh", methods=["POST"])
@login_required
@admin_required
def refresh_tracker():
    ensure_tracker_schema()
    refresh_tracker_sheet_results()
    flash("Tracker results refreshed.", "success")
    return redirect(url_for("daily_tracker"))


if __name__ == "__main__":
    ensure_db_columns()
    ensure_tracker_schema()

    import os
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        threading.Thread(target=run_scheduler, daemon=True).start()

    app.run(debug=True)

# Historical Report Generator Feature
HISTORICAL_REPORT_FEATURE = True
